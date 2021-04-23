# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
#import io
import sys
import configparser
#import time
import shutil
import openpyxl                       # Для .xlsx
# import xlrd                         # для .xls
from   price_tools import getCellXlsx, quoted, dump_cell, currencyType, subInParentheses
import requests



def download( ):
    retCode     = False
    filename_new= 'new_profdisplay.xlsx'
    filename_old= 'old_profdisplay.xlsx'
    url_file    = 'http://displays-price.ru/price.xlsx'
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.0; rv:14.0) Gecko/20100101 Firefox/14.0.1',
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
               'Accept-Language':'ru-ru,ru;q=0.8,en-us;q=0.5,en;q=0.3',
               'Accept-Encoding':'gzip, deflate',
               'Connection':'keep-alive',
               'DNT':'1'
              }    
    try:
        s = requests.Session()
        #r = s.get(url_lk,  headers = headers) 
        #print(r.text)                         # посмотреть исходный код страницы
        '''
        page = lxml.html.fromstring(r.text)
        for ff in page.forms:                 # посмотреть список форм и их поля
            print(ff.fields.keys())
        form = page.forms[0]
        form.fields['login'] = login
        form.fields['pass'] = password
        r = s.post(url_lk+ form.action, data=form.form_values())
        print('       ==================================================')
        #print('<<<',r.text,'>>>')

        log.debug('Авторизация на %s   --- code=%d', url_lk, r.status_code)
        '''
        r = s.get(url_file)
        log.debug('Загрузка файла %16d bytes   --- code=%d', len(r.content), r.status_code)
        retCode = True
        '''
        s = requests.Session()
        r = s.get(url_lk, auth=(login,password))  # ,headers = headers (И без него сработало, но где-то может понадобиться)
        page = lxml.html.fromstring(r.text)
        # data = {'USER_LOGIN':login, 'USER_PASSWORD':password})
        log.debug('Авторизация на %s   --- code=%d', url_lk, r.status_code)
        r = s.get(url_file)
        log.debug('Загрузка файла %24d bytes   --- code=%d', len(r.content), r.status_code)
        retCode = True
        '''
    except Exception as e:
        log.debug('Exception: <' + str(e) + '>')

    if os.path.exists( filename_new) and os.path.exists( filename_old): 
        os.remove( filename_old)
        os.rename( filename_new, filename_old)
    if os.path.exists( filename_new) :
        os.rename( filename_new, filename_old)
    f2 = open(filename_new, 'wb')                                  #Теперь записываем файл
    f2.write(r.content)
    f2.close()
    return retCode



def convert_sheet( book, sheetName, confName):
    csvFName = confName.replace('cfg','csv')
    if not os.path.exists( confName ) :
        log.error( 'Нет файла конфигурации '+confName)
        return
    # Прочитать конфигурацию из файла
    in_columns_j, out_columns = config_read( confName )
    sh = book[sheetName]                                     # xlsx
    ssss = []
    subgrp = ''
    for i in range(sh.min_row, sh.max_row) :
        i_last = i
        try:
            ccc =sh.cell(row=i, column=in_columns_j['категория'] )
            cc2 =sh.cell(row=i, column= 11 )
            '''
            if  (sh.cell(row=i, column=in_columns_j['продажа'] ).value== None)        or \
                (confName=='cfg_поnec.cfg' and sh.cell(row=i, column=3 ).value==None) or \
                (confName=='cfg_проекторыpanasonic.cfg' and cc2.value==None)          or \
                (confName=='cfg_проект_акс_panas.cfg'   and cc2.value!=None)       :  # ненужная строка
                    continue
            '''
            if not(          # Не строка нужной таблицы
                (('nonblank' not in in_columns_j.keys()) or (sh.cell(row=i, column=in_columns_j['nonblank']).value!=None)) and
                (('blank' not in in_columns_j.keys()) or (sh.cell(row=i, column=in_columns_j['blank']).value==None)
                                                      or (sh.cell(row=i, column=in_columns_j['blank']).value==''))):
                continue
            elif ccc.value!=None and ((ccc.value[0:9]=='Категория') or (ccc.value=='SOFTWARE SOLUTIONS')):  # Заголовок таблицы
                continue
            else :                                                                    # Информационная строка или подгруппа
                impValues = getXlsxString(sh, i, in_columns_j)
                if (impValues['закупка'] == '0.1') or (impValues['закупка'] == '0'):
                    impValues['закупка'] = impValues['продажа']
                if confName == 'cfg_philips.cfg' and impValues['валюта_по_формату'] != 'USD':
                    continue
                elif confName == 'cfg_philips_rur.cfg' and impValues['валюта_по_формату'] != 'RUB':
                    continue
                if  confName[-7:-4] =='aks' :                                         # Для аксессуарров наследуем подгруппу
                    if  impValues['категория'] == '' :
                        impValues['категория'] = subgrp
                    else :
                        subgrp = impValues['категория']
                sss = []                                                              # формируемая строка для вывода в файл
                for outColName in out_columns.keys() :
                    shablon = out_columns[outColName]
                    for key in impValues.keys():
                        if shablon.find(key) >= 0 :
                            shablon = shablon.replace(key, impValues[key])
                    if   (outColName == 'описание') and ('тип_сенсора' in impValues) :
                       shablon = appendSensor( shablon, impValues)
                    elif (outColName == 'закупка') and ('*' in shablon) :
                       p = shablon.find("*")
                       vvv1 = float(shablon[:p])
                       vvv2 = float(shablon[p+1:])
                       shablon = str(round(vvv1 * vvv2, 2))
                    sss.append( quoted( shablon))
                ssss.append(','.join(sss))
                    
        except Exception as e:
            log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'<' + '>' )
            raise e

    log.info('Обработано ' +str(i_last)+ ' строк.')
    f2 = open( csvFName, 'w', encoding='cp1251')
    strHeader = ','.join( out_columns.keys() ) + ','
    f2.write( strHeader + '\n' )
    data = ',\n'.join(ssss) +','
    bbbb = data.encode(encoding='cp1251', errors='replace')
    data = bbbb.decode(encoding='cp1251')
    f2.write(data)
    f2.close()



def appendSensor( shablon, impValues):
    ss = impValues['тип_сенсора']
    if ss != 'нет' :  shablon = shablon + '\nтип сенсора: ' + ss
    ss = impValues['количество_точек_касания']
    if ss != 'нет' :  shablon = shablon + '\nколичество точек касания: ' + ss
    return shablon



def currencyType( row, col, sheet ):
    '''
    Функция анализирует "формат ячейки" таблицы excel, является ли он "денежным"
    и какая валюта указана в этом формате.
    Распознаются не все валюты и способы их описания.
    '''
    c = sheet.cell( row=row, column=col )
    '''                                                  # -- для XLS
    xf = sheet.book.xf_list[c.xf_index]
    fmt_obj = sheet.book.format_map[xf.format_key]
    fmt_str = fmt_obj.format_str
    '''                                                  # -- для XLSX
    fmt_str = c.number_format

    if 'р' in fmt_str:
        val = 'RUB'
    elif '\xa3' in fmt_str:
        val = 'GBP'
    elif chr(8364) in fmt_str:
        val = 'EUR'
    elif (fmt_str.find('USD')>=0) or (fmt_str.find('[$$')>=0) :
        val = 'USD'
    else:
        val = ''
    return val



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена') :
            if getCellXlsx(row=i, col=j, isDigit='N', sheet=sh).find('по запросу') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def config_read( cfgFName ):
    log.debug('Reading config ' + cfgFName )
    
    config = configparser.ConfigParser()
    if os.path.exists(cfgFName):     config.read( cfgFName, encoding='utf-8')
    else : log.debug('Не найден файл конфигурации.')

    # в разделе [cols_in] находится список интересующих нас колонок и номера столбцов исходного файла
    in_columns_names = config.options('cols_in')
    in_columns_j = {}
    for vName in in_columns_names :
        if ('' != config.get('cols_in', vName)) :
            in_columns_j[vName] = config.getint('cols_in', vName) 
    
    # По разделу [cols_out] формируем перечень выводимых колонок и строку заголовка результирующего CSV файла
    out_columns_names = config.options('cols_out')
    out_columns = {}
    for vName in out_columns_names :
        if ('' != config.get('cols_out', vName)) :
            out_columns[vName] = config.get('cols_out', vName) 

    return in_columns_j, out_columns



def convert2csv( dealerName ):
    fileNameIn = 'new_'+dealerName+'.xlsx'
    book = openpyxl.load_workbook(filename = fileNameIn, read_only=False, keep_vba=False, data_only=False)
#   book = xlrd.open_workbook( fileNameIn.encode('cp1251'), formatting_info=True)
#    sheetNames = book.sheetnames() #  .get_sheet_names()
    for sheetName in book.sheetnames :                                # Организую цикл по страницам
        log.info('-------------------  '+sheetName +'  ----------')
        confName = ('cfg_'+sheetName.replace(' ','').replace('.','')+'.cfg').lower()
        if   sheetName.upper() == 'SAMSUNG'       : convert_sheet( book, sheetName, confName)
        elif sheetName.upper() == 'LG'            : convert_sheet( book, sheetName, confName)
        elif sheetName.upper() == 'NEC'           : 
                                                    convert_sheet( book, sheetName, confName)
                                                    convert_sheet( book, sheetName, 'cfg_nec_aks.cfg')
                                                    convert_sheet( book, sheetName, 'cfg_nec_soft.cfg')
        elif sheetName.upper() == 'BENQ'          : convert_sheet( book, sheetName, confName)
        elif sheetName.upper() == 'SHARP'         : convert_sheet( book, sheetName, confName)
        elif sheetName.upper() == 'IIYAMA'        : convert_sheet( book, sheetName, confName)
        elif sheetName.upper() == 'PHILIPS'       :
                                                    convert_sheet( book, sheetName, confName)
                                                    convert_sheet( book, sheetName, 'cfg_philips_rur.cfg')
        elif sheetName.upper() == 'VIEWSONIC'     : convert_sheet( book, sheetName, confName)
        elif sheetName.upper() == 'PANASONIC'     : convert_sheet( book, sheetName, confName)
        elif sheetName.upper() == 'ПРОЕКТОРЫ PANASONIC':              # больше не предоставляют
                                                    convert_sheet( book, sheetName, confName)
                                                    convert_sheet( book, sheetName, 'cfg_проекторыpanasonic_aks.cfg')
        else : log.debug('Не конвертируем лист '+sheetName )



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def main( dealerName):
    make_loger()
    log.info('         '+dealerName )
    rc_download = download()
    if rc_download==True or is_file_fresh( filename_new, 3):
        convert2csv( dealerName )



if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( 'profdisplay')
